"""Exports Excel et PDF des états financiers."""

import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.core.periods import short_label
from apps.students.models import Student, StudentStatus

from .services import bilan, encaissements

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Le franc CFA s'écrit sans décimale et avec une espace comme séparateur de milliers.
MONEY_FORMAT = "# ##0"


def _xlsx_response(workbook):
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return HttpResponse(stream.read(), content_type=XLSX_MIME)


def _write_header(sheet, row, labels, widths=None):
    for index, label in enumerate(labels, start=1):
        cell = sheet.cell(row=row, column=index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for index, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_series(sheet, row, label, values, total, bold=False):
    cell = sheet.cell(row=row, column=1, value=label)
    cell.border = BORDER
    if bold:
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
    for offset, value in enumerate(values, start=2):
        cell = sheet.cell(row=row, column=offset, value=value)
        cell.number_format = MONEY_FORMAT
        cell.border = BORDER
        if bold:
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
    cell = sheet.cell(row=row, column=len(values) + 2, value=total)
    cell.number_format = MONEY_FORMAT
    cell.font = TOTAL_FONT
    cell.border = BORDER
    if bold:
        cell.fill = TOTAL_FILL


def encais_xlsx(year, request):
    data = encaissements(year)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ENCAIS"

    periods = [short_label(p["date"]) for p in data["periods"]]
    sheet["A1"] = f"Synthèse des encaissements — {data['year']}"
    sheet["A1"].font = Font(bold=True, size=14)

    _write_header(
        sheet, 3, ["INTITULE", "EFFECTIF"] + periods + ["TOTAL"],
        widths=[34, 10] + [13] * len(periods) + [15],
    )

    row = 4
    for entry in data["classes"]:
        sheet.cell(row=row, column=1, value=f"Inscription reçue — {entry['classroom']}").border = BORDER
        sheet.cell(row=row, column=2, value=entry["headcount"]).border = BORDER
        for offset, value in enumerate(entry["registration"]["values"], start=3):
            cell = sheet.cell(row=row, column=offset, value=value)
            cell.number_format = MONEY_FORMAT
            cell.border = BORDER
        cell = sheet.cell(row=row, column=len(periods) + 3, value=entry["registration"]["total"])
        cell.number_format = MONEY_FORMAT
        cell.border = BORDER
        row += 1

    _write_series(
        sheet, row, "Total inscription reçue",
        [0] + data["registration_total"]["values"],
        data["registration_total"]["total"], bold=True,
    )
    row += 2

    for entry in data["classes"]:
        sheet.cell(row=row, column=1, value=f"Mensualité reçue — {entry['classroom']}").border = BORDER
        sheet.cell(row=row, column=2, value=entry["headcount"]).border = BORDER
        for offset, value in enumerate(entry["tuition"]["values"], start=3):
            cell = sheet.cell(row=row, column=offset, value=value)
            cell.number_format = MONEY_FORMAT
            cell.border = BORDER
        cell = sheet.cell(row=row, column=len(periods) + 3, value=entry["tuition"]["total"])
        cell.number_format = MONEY_FORMAT
        cell.border = BORDER
        row += 1

    _write_series(
        sheet, row, "Total mensualité reçue",
        [0] + data["tuition_total"]["values"],
        data["tuition_total"]["total"], bold=True,
    )

    sheet.freeze_panes = "C4"
    return _xlsx_response(workbook)


def bilan_xlsx(year, request):
    data = bilan(year)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rapport Bilan"

    periods = [short_label(p["date"]) for p in data["periods"]]
    sheet["A1"] = f"Rapport bilan — année scolaire {data['year']}"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"] = f"Établissement : {request.user.school.name}"

    _write_header(
        sheet, 4, ["RUBRIQUE"] + periods + ["TOTAL"],
        widths=[46] + [13] * len(periods) + [16],
    )

    row = 5
    for entry in data["resources"]:
        _write_series(sheet, row, entry["label"], entry["values"], entry["total"])
        row += 1
    _write_series(
        sheet, row, "TOTAL RESSOURCE",
        data["total_resources"]["values"], data["total_resources"]["total"], bold=True,
    )
    row += 2

    for entry in data["charges"]:
        _write_series(sheet, row, entry["label"], entry["values"], entry["total"])
        row += 1
    _write_series(
        sheet, row, "TOTAL CHARGE",
        data["total_charges"]["values"], data["total_charges"]["total"], bold=True,
    )
    row += 2

    _write_series(sheet, row, "EXCÉDENT BRUT D'EXPLOITATION (EBE)",
                  data["ebe"]["values"], data["ebe"]["total"], bold=True)
    row += 1
    _write_series(sheet, row, "SOLDE CUMULE",
                  data["cumulative_balance"]["values"], data["current_balance"], bold=True)
    row += 2

    sheet.cell(row=row, column=1, value="LE SOLDE DU COMPTE À CE JOUR EST :").font = TOTAL_FONT
    cell = sheet.cell(row=row, column=2, value=data["current_balance"])
    cell.number_format = MONEY_FORMAT
    cell.font = TOTAL_FONT
    row += 3

    # Effectifs et chiffre d'affaires par classe.
    _write_header(sheet, row, ["EFFECTIF DES CLASSES ET REVENUS"] +
                  [c["classroom"] for c in data["headcount_by_class"]] + ["TOTAL"])
    row += 1
    sheet.cell(row=row, column=1, value="EFFECTIF TOTAL DES ELEVES").border = BORDER
    for offset, entry in enumerate(data["headcount_by_class"], start=2):
        sheet.cell(row=row, column=offset, value=entry["headcount"]).border = BORDER
    sheet.cell(row=row, column=len(data["headcount_by_class"]) + 2,
               value=data["headcount_total"]).font = TOTAL_FONT
    row += 1
    sheet.cell(row=row, column=1, value="CHIFFRE D'AFFAIRE PAR CLASSE").border = BORDER
    for offset, entry in enumerate(data["headcount_by_class"], start=2):
        cell = sheet.cell(row=row, column=offset, value=entry["revenue"])
        cell.number_format = MONEY_FORMAT
        cell.border = BORDER
    cell = sheet.cell(row=row, column=len(data["headcount_by_class"]) + 2,
                      value=data["revenue_total"])
    cell.number_format = MONEY_FORMAT
    cell.font = TOTAL_FONT

    sheet.freeze_panes = "B5"
    return _xlsx_response(workbook)


def students_xlsx(year, request):
    """Liste nominative des élèves, par classe."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Élèves"

    _write_header(
        sheet, 1,
        ["N°", "Prénom", "Nom", "Classe", "Date de naissance", "Sexe",
         "Parent / tuteur", "Téléphone", "Statut"],
        widths=[6, 18, 18, 12, 16, 8, 24, 16, 14],
    )

    students = Student.objects.select_related("classroom").order_by(
        "classroom__order", "last_name", "first_name"
    )
    for index, student in enumerate(students, start=1):
        sheet.append([
            index,
            student.first_name,
            student.last_name,
            student.classroom.name,
            student.date_of_birth,
            student.get_sex_display() if student.sex else "",
            student.parent_name,
            student.parent_phone,
            student.get_status_display(),
        ])

    sheet.freeze_panes = "A2"
    return _xlsx_response(workbook)


def bilan_pdf(year, request):
    """Rapport bilan en PDF, horodaté et attribué à son auteur.

    L'horodatage et l'identité du générateur sont en pied de page : ce document est
    destiné à être présenté à un conseil d'administration ou à une tutelle, et doit
    donc porter la trace de qui l'a produit et quand.
    """
    from django.utils import timezone

    data = bilan(year)
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=16 * mm,
        title=f"Rapport bilan {data['year']}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=15, spaceAfter=4)
    meta_style = ParagraphStyle("m", parent=styles["Normal"], fontSize=8,
                                textColor=colors.HexColor("#555555"))

    generated_at = timezone.localtime().strftime("%d/%m/%Y à %H:%M")
    author = request.user.get_full_name() or request.user.email

    story = [
        Paragraph(f"Rapport bilan — année scolaire {data['year']}", title_style),
        Paragraph(request.user.school.name, styles["Normal"]),
        Spacer(1, 6),
    ]

    periods = [short_label(p["date"]) for p in data["periods"]]

    def money(value):
        return f"{value:,}".replace(",", " ") if value else "—"

    rows = [["RUBRIQUE"] + periods + ["TOTAL"]]
    section_rows = []

    for entry in data["resources"]:
        rows.append([entry["label"]] + [money(v) for v in entry["values"]] + [money(entry["total"])])
    section_rows.append(len(rows) - 1)
    rows.append(["TOTAL RESSOURCE"] + [money(v) for v in data["total_resources"]["values"]]
                + [money(data["total_resources"]["total"])])
    total_resource_row = len(rows) - 1

    for entry in data["charges"]:
        rows.append([entry["label"]] + [money(v) for v in entry["values"]] + [money(entry["total"])])
    rows.append(["TOTAL CHARGE"] + [money(v) for v in data["total_charges"]["values"]]
                + [money(data["total_charges"]["total"])])
    total_charge_row = len(rows) - 1

    rows.append(["EXCÉDENT BRUT D'EXPLOITATION"] + [money(v) for v in data["ebe"]["values"]]
                + [money(data["ebe"]["total"])])
    ebe_row = len(rows) - 1
    rows.append(["SOLDE CUMULE"] + [money(v) for v in data["cumulative_balance"]["values"]]
                + [money(data["current_balance"])])
    balance_row = len(rows) - 1

    table = Table(rows, repeatRows=1, colWidths=[68 * mm] + [14 * mm] * len(periods) + [20 * mm])
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.2),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FB")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    for index in (total_resource_row, total_charge_row, ebe_row, balance_row):
        style += [
            ("BACKGROUND", (0, index), (-1, index), colors.HexColor("#D9E2F3")),
            ("FONTNAME", (0, index), (-1, index), "Helvetica-Bold"),
        ]
    table.setStyle(TableStyle(style))
    story.append(table)

    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"<b>Solde du compte à ce jour : {money(data['current_balance'])} "
        f"{request.user.school.currency}</b>", styles["Normal"]))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"Effectif total : {data['headcount_total']} élèves — "
        f"chiffre d'affaires : {money(data['revenue_total'])} {request.user.school.currency}",
        styles["Normal"]))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawString(
            12 * mm, 8 * mm,
            f"Généré le {generated_at} par {author} — MonÉcole",
        )
        canvas.drawRightString(
            landscape(A4)[0] - 12 * mm, 8 * mm, f"Page {doc.page}"
        )
        canvas.restoreState()

    document.build(story, onFirstPage=footer, onLaterPages=footer)
    buffer.seek(0)
    return HttpResponse(buffer.read(), content_type="application/pdf")
