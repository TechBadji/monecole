"""Lecture des classeurs Excel.

Deux dispositions sont acceptées, parce que les écoles arrivent avec l'une ou
l'autre :

1. **Tableau simple** — une feuille, une ligne d'en-tête, une ligne par
   enregistrement. C'est le modèle que MonÉcole propose au téléchargement, en
   `.xlsx` comme en CSV.

2. **Classeur de gestion** — la structure décrite dans `docs/modele-excel.md` :
   un onglet par classe (GARDERIE, PS, … CM2), en-têtes en ligne 8, élèves à
   partir de la ligne 9, et les blocs d'inscription et de mensualités en
   colonnes. C'est le format réel dans lequel vivent les données des écoles, et
   pouvoir l'ingérer directement épargne dix conversions manuelles.

La disposition est **détectée**, pas demandée : une école qui dépose son classeur
ne sait pas dans quelle catégorie il tombe.
"""

import datetime
import io

from openpyxl import load_workbook

from .imports import normalize_header, parse_amount, parse_date

# Onglets de classe du classeur de gestion, dans l'ordre pédagogique.
CLASS_SHEETS = [
    "GARDERIE", "PS", "MS", "GS", "CI", "CP", "CE1", "CE2", "CM1", "CM2",
]

# Disposition des onglets de classe (voir docs/modele-excel.md).
HEADER_ROW = 8
FIRST_STUDENT_ROW = 9
COL_FIRST_NAME = 4      # D
COL_LAST_NAME = 5       # E
COL_BIRTH_DATE = 7      # G
COL_REGISTRATION_PAID = 8   # H
COL_REGISTRATION = 9        # I
COL_UNIFORM = 10            # J
COL_INSURANCE = 11          # K
COL_APE = 12                # L
COL_TUITION_FIRST = 13      # M — neuf mois, d'octobre à juin
TUITION_MONTHS = 9

# Au-delà, on considère la feuille terminée : le classeur d'origine prévoit des
# lignes vides jusqu'à la 530e, qu'il ne sert à rien de parcourir.
MAX_BLANK_ROWS = 25


class WorkbookError(Exception):
    """Le classeur ne peut pas être lu."""


def open_workbook(raw: bytes):
    try:
        return load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as error:  # noqa: BLE001 — openpyxl lève des types variés
        raise WorkbookError(
            "Fichier Excel illisible. Vérifiez qu'il s'agit bien d'un .xlsx "
            "(et non d'un .xls ancien format, à réenregistrer depuis Excel)."
        ) from error


def detect_layout(workbook):
    """« management » si le classeur porte les onglets de classe, sinon « table »."""
    names = {sheet.upper().strip() for sheet in workbook.sheetnames}
    matched = [name for name in CLASS_SHEETS if name in names]
    # Trois onglets de classe suffisent à trancher : une école peut n'avoir que
    # l'élémentaire, ou avoir renommé un onglet.
    return "management" if len(matched) >= 3 else "table"


# --------------------------------------------------------------------------- #
# Disposition 1 — tableau simple                                               #
# --------------------------------------------------------------------------- #


def read_table(workbook, sheet_name=None):
    """Lit une feuille tabulaire et retourne des lignes normalisées."""
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        raise WorkbookError("Feuille vide.")

    columns = [normalize_header(str(value or "")) for value in header]
    results = []
    for values in rows:
        if all(value in (None, "") for value in values):
            continue
        results.append(
            {
                column: _stringify(value)
                for column, value in zip(columns, values)
                if column
            }
        )
    return results


def _stringify(value):
    """Ramène une cellule à du texte, en préservant les dates et les entiers.

    Excel rend les dates comme des `datetime` et les montants comme des `float` :
    les convertir naïvement produirait « 15000.0 » et « 2018-05-12 00:00:00 ».
    """
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


# --------------------------------------------------------------------------- #
# Disposition 2 — classeur de gestion                                          #
# --------------------------------------------------------------------------- #


def read_management_workbook(workbook, year):
    """Extrait élèves, inscriptions et mensualités des onglets de classe.

    Retourne `(records, report)` où chaque enregistrement porte l'élève et ses
    encaissements. Les anomalies sont collectées plutôt que levées : un classeur
    réel comporte toujours des lignes bancales, et refuser tout le fichier pour
    une cellule douteuse serait inexploitable.
    """
    records = []
    warnings = []
    tuition_periods = year.tuition_month_ends[:TUITION_MONTHS]

    available = {sheet.upper().strip(): sheet for sheet in workbook.sheetnames}

    for class_name in CLASS_SHEETS:
        actual = available.get(class_name)
        if actual is None:
            continue
        sheet = workbook[actual]

        blank_streak = 0
        for row_index in range(FIRST_STUDENT_ROW, sheet.max_row + 1):
            first_name = _stringify(sheet.cell(row_index, COL_FIRST_NAME).value)
            last_name = _stringify(sheet.cell(row_index, COL_LAST_NAME).value)

            if not first_name and not last_name:
                blank_streak += 1
                if blank_streak >= MAX_BLANK_ROWS:
                    break
                continue
            blank_streak = 0

            if not first_name or not last_name:
                warnings.append(
                    f"{class_name} ligne {row_index} : prénom ou nom manquant, ignorée."
                )
                continue

            birth_date = None
            raw_birth = sheet.cell(row_index, COL_BIRTH_DATE).value
            if raw_birth:
                try:
                    birth_date = parse_date(_stringify(raw_birth))
                except ValueError:
                    warnings.append(
                        f"{class_name} ligne {row_index} : date de naissance "
                        f"« {raw_birth} » illisible, laissée vide."
                    )

            registration = _amount(sheet, row_index, COL_REGISTRATION, class_name,
                                   row_index, warnings)
            tuition = {}
            for offset, period in enumerate(tuition_periods):
                value = _amount(
                    sheet, row_index, COL_TUITION_FIRST + offset,
                    class_name, row_index, warnings,
                )
                if value:
                    tuition[period] = value

            records.append(
                {
                    "class_name": class_name,
                    "row": row_index,
                    "first_name": first_name,
                    "last_name": last_name,
                    "date_of_birth": birth_date,
                    "registration_amount": registration,
                    "registration_paid": bool(
                        sheet.cell(row_index, COL_REGISTRATION_PAID).value
                    ),
                    "uniform_amount": _amount(sheet, row_index, COL_UNIFORM,
                                              class_name, row_index, warnings),
                    "insurance_amount": _amount(sheet, row_index, COL_INSURANCE,
                                                class_name, row_index, warnings),
                    "ape_amount": _amount(sheet, row_index, COL_APE,
                                          class_name, row_index, warnings),
                    "tuition": tuition,
                }
            )

    return records, warnings


def _amount(sheet, row, column, class_name, row_index, warnings):
    raw = sheet.cell(row, column).value
    if raw in (None, ""):
        return 0
    try:
        return parse_amount(_stringify(raw))
    except ValueError:
        warnings.append(
            f"{class_name} ligne {row_index} : montant « {raw} » illisible, compté zéro."
        )
        return 0


def summarize(records):
    """Aperçu du classeur, avant toute écriture."""
    by_class = {}
    for record in records:
        entry = by_class.setdefault(
            record["class_name"],
            {"classroom": record["class_name"], "students": 0,
             "registration": 0, "tuition": 0},
        )
        entry["students"] += 1
        entry["registration"] += record["registration_amount"]
        entry["tuition"] += sum(record["tuition"].values())

    classes = list(by_class.values())
    return {
        "students": len(records),
        "classes": classes,
        "total_registration": sum(entry["registration"] for entry in classes),
        "total_tuition": sum(entry["tuition"] for entry in classes),
    }
