"""Import d'un classeur Excel de gestion.

Le classeur fourni par l'école est le gabarit V0, vide de données. Les tests
construisent donc un classeur **à sa structure exacte** — onglets de classe,
en-têtes ligne 8, élèves ligne 9, blocs d'inscription et de mensualités aux mêmes
colonnes — et vérifient que la reprise en tire les bons montants.
"""

import datetime
import io

from django.test import TestCase
from openpyxl import Workbook
from rest_framework.test import APIClient

from apps.core.models import Role
from apps.core.tenancy import tenant_context
from apps.core.tests.factories import make_classroom, make_school, make_user, make_year
from apps.core.workbook_import import detect_layout, open_workbook
from apps.students.models import Enrollment, MonthlyPayment, Student


def management_workbook(rows_by_class, tuition_start=13):
    """Classeur à la disposition du classeur de gestion réel."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for class_name, rows in rows_by_class.items():
        sheet = workbook.create_sheet(class_name)
        # Ligne 8 : en-têtes, comme dans le classeur d'origine.
        sheet.cell(8, 4, "Prénom")
        sheet.cell(8, 5, "Nom")
        sheet.cell(8, 7, "Date de naissance")
        sheet.cell(8, 9, "Montant inscription")

        for offset, row in enumerate(rows):
            index = 9 + offset
            sheet.cell(index, 4, row["first_name"])
            sheet.cell(index, 5, row["last_name"])
            if row.get("birth"):
                sheet.cell(index, 7, row["birth"])
            sheet.cell(index, 8, row.get("paid", 1))
            sheet.cell(index, 9, row.get("registration", 0))
            sheet.cell(index, 10, row.get("uniform", 0))
            sheet.cell(index, 11, row.get("insurance", 0))
            sheet.cell(index, 12, row.get("ape", 0))
            for month, amount in enumerate(row.get("tuition", [])):
                sheet.cell(index, tuition_start + month, amount)

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


def simple_table(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Prénom", "Nom", "Date de naissance", "Sexe", "Classe"])
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


class LayoutDetectionTests(TestCase):
    def test_management_layout_is_detected(self):
        raw = management_workbook({"CI": [], "CP": [], "CE1": []})
        self.assertEqual(detect_layout(open_workbook(raw)), "management")

    def test_two_class_sheets_are_not_enough(self):
        """Un seuil de trois onglets évite de confondre avec un tableau nommé « CP »."""
        raw = management_workbook({"CI": [], "CP": []})
        self.assertEqual(detect_layout(open_workbook(raw)), "table")

    def test_simple_table_is_detected(self):
        raw = simple_table([["Awa", "Diop", "12/05/2018", "F", "CP"]])
        self.assertEqual(detect_layout(open_workbook(raw)), "table")

    def test_corrupt_file_is_reported_clearly(self):
        from apps.core.workbook_import import WorkbookError

        with self.assertRaises(WorkbookError) as caught:
            open_workbook(b"ceci n'est pas un classeur")
        self.assertIn(".xlsx", str(caught.exception))


class WorkbookImportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.year = make_year(cls.school)
        cls.cp = make_classroom(cls.school, "CP", order=5)
        cls.admin = make_user(cls.school, Role.ADMIN, "admin@test.sn")

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def upload(self, raw, dry_run=True):
        return self.client.post(
            "/api/imports/workbook/",
            {
                "file": io.BytesIO(raw),
                "dry_run": "true" if dry_run else "false",
            },
            format="multipart",
        )

    ROWS = {
        "CP": [
            {
                "first_name": "Awa", "last_name": "Diop", "birth": "12/05/2018",
                "registration": 25_000, "uniform": 12_000, "insurance": 3_000,
                "ape": 5_000, "tuition": [15_000, 15_000, 15_000],
            },
            {
                "first_name": "Moussa", "last_name": "Fall", "birth": "03/09/2017",
                "registration": 25_000, "tuition": [15_000, 0, 15_000],
            },
        ],
        "CE1": [
            {
                "first_name": "Bineta", "last_name": "Sow",
                "registration": 25_000, "tuition": [16_000],
            },
        ],
        "CM2": [],
    }

    def test_dry_run_reports_without_writing(self):
        response = self.upload(management_workbook(self.ROWS))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["layout"], "management")
        self.assertFalse(response.data["applied"])
        self.assertEqual(response.data["students"], 3)
        # 3 × 25 000 d'inscription.
        self.assertEqual(response.data["total_registration"], 75_000)
        # (15 000 × 3) + (15 000 + 15 000) + 16 000 = 91 000
        self.assertEqual(response.data["total_tuition"], 91_000)

        with tenant_context(self.school):
            self.assertEqual(Student.objects.count(), 0)

    def test_applying_creates_students_enrollments_and_payments(self):
        response = self.upload(management_workbook(self.ROWS), dry_run=False)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["applied"])
        self.assertEqual(response.data["created"], 3)

        with tenant_context(self.school):
            self.assertEqual(Student.objects.count(), 3)
            self.assertEqual(Enrollment.objects.count(), 3)
            # 3 + 2 mensualités non nulles + 1 = 6 ; le zéro de novembre est ignoré.
            self.assertEqual(MonthlyPayment.objects.count(), 6)

            awa = Student.objects.get(first_name="Awa")
            self.assertEqual(awa.date_of_birth, datetime.date(2018, 5, 12))
            self.assertEqual(awa.classroom.name, "CP")
            self.assertRegex(awa.matricule, r"^M\d{4}$")

            enrollment = Enrollment.objects.get(student=awa)
            self.assertEqual(enrollment.registration_amount, 25_000)
            self.assertEqual(enrollment.uniform_amount, 12_000)

    def test_missing_classes_are_created_at_the_right_rank(self):
        """Le CE1 n'existe pas en base : il doit être créé après le CP."""
        self.upload(management_workbook(self.ROWS), dry_run=False)
        with tenant_context(self.school):
            from apps.students.models import ClassRoom

            ce1 = ClassRoom.objects.get(name="CE1")
            self.assertEqual(ce1.order, 6)
            self.assertEqual(ce1.level, "PRIMARY")

    def test_reimport_keeps_the_matricule(self):
        """Une reprise ne doit jamais renuméroter un élève déjà inscrit."""
        self.upload(management_workbook(self.ROWS), dry_run=False)
        with tenant_context(self.school):
            before = Student.objects.get(first_name="Awa").matricule

        response = self.upload(management_workbook(self.ROWS), dry_run=False)
        self.assertEqual(response.data["updated"], 3)
        self.assertEqual(response.data["created"], 0)

        with tenant_context(self.school):
            self.assertEqual(Student.objects.get(first_name="Awa").matricule, before)
            self.assertEqual(Student.objects.count(), 3)

    def test_unreadable_amount_is_warned_not_fatal(self):
        """Un classeur réel comporte toujours des cellules bancales."""
        rows = {"CI": [], "CP": [
            {"first_name": "Awa", "last_name": "Diop", "registration": "à revoir"},
        ], "CE1": [], "CM2": []}
        response = self.upload(management_workbook(rows))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data["ok"])
        self.assertGreaterEqual(response.data["warning_count"], 1)
        self.assertIn("illisible", str(response.data["warnings"]))

    def test_row_without_a_name_is_skipped_with_a_warning(self):
        rows = {"CI": [], "CP": [
            {"first_name": "Awa", "last_name": ""},
            {"first_name": "Moussa", "last_name": "Fall", "registration": 25_000},
        ], "CE1": [], "CM2": []}
        response = self.upload(management_workbook(rows))
        self.assertEqual(response.data["students"], 1)
        self.assertIn("manquant", str(response.data["warnings"]))

    def test_empty_workbook_is_refused_with_guidance(self):
        """Le gabarit V0 de l'école est vide : le message doit orienter."""
        response = self.upload(management_workbook({"CI": [], "CP": [], "CE1": []}))
        self.assertEqual(response.status_code, 400)
        self.assertIn("ligne 9", str(response.data))

    def test_simple_table_goes_through_the_csv_pipeline(self):
        raw = simple_table([
            ["Awa", "Diop", "12/05/2018", "F", "CP"],
            ["Moussa", "Fall", "03/09/2017", "M", "CP"],
        ])
        response = self.client.post(
            "/api/imports/workbook/",
            {"file": io.BytesIO(raw), "kind": "students", "dry_run": "false"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["layout"], "table")
        self.assertEqual(response.data["created"], 2)

    def test_excel_dates_and_floats_survive_the_table_pipeline(self):
        """Excel rend « 12/05/2018 » en datetime et « 15000 » en float."""
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Prénom", "Nom", "Date de naissance", "Classe"])
        sheet.append(["Awa", "Diop", datetime.datetime(2018, 5, 12), "CP"])
        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = self.client.post(
            "/api/imports/workbook/",
            {"file": io.BytesIO(stream.read()), "kind": "students", "dry_run": "false"},
            format="multipart",
        )
        self.assertEqual(response.status_code, 200)
        with tenant_context(self.school):
            self.assertEqual(
                Student.objects.get(first_name="Awa").date_of_birth,
                datetime.date(2018, 5, 12),
            )

    def test_accountant_cannot_import_a_workbook(self):
        accountant = make_user(self.school, Role.ACCOUNTANT, "compta@test.sn")
        client = APIClient()
        client.force_authenticate(accountant)
        response = client.post(
            "/api/imports/workbook/",
            {"file": io.BytesIO(management_workbook(self.ROWS))},
            format="multipart",
        )
        self.assertEqual(response.status_code, 403)

    def test_formats_are_documented(self):
        response = self.client.get("/api/imports/workbook/")
        self.assertEqual(response.status_code, 200)
        keys = {entry["key"] for entry in response.data["layouts"]}
        self.assertEqual(keys, {"management", "table"})
