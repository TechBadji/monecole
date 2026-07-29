"""Journal d'audit.

Critère d'acceptation du cahier des charges : « toute modification ou suppression
d'un paiement ou d'une dépense est tracée dans l'AuditLog avec l'utilisateur
responsable ».
"""

from django.test import TestCase
from rest_framework.test import APIClient

from apps.core.models import AuditLog, Role
from apps.core.tenancy import tenant_context
from apps.students.models import MonthlyPayment

from .factories import (
    make_category,
    make_classroom,
    make_fee_schedule,
    make_school,
    make_student,
    make_user,
    make_year,
)


class AuditImmutabilityTests(TestCase):
    """Un journal modifiable ne prouve rien."""

    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.user = make_user(cls.school, Role.ADMIN, "admin@test.sn")

    def make_entry(self):
        return AuditLog.objects.create(
            school=self.school, user=self.user, user_label=str(self.user),
            action=AuditLog.Action.CREATE, entity="Expense", entity_id="1",
        )

    def test_an_entry_cannot_be_modified(self):
        entry = self.make_entry()
        entry.action = AuditLog.Action.DELETE
        with self.assertRaises(PermissionError):
            entry.save()

    def test_an_entry_cannot_be_deleted(self):
        entry = self.make_entry()
        with self.assertRaises(PermissionError):
            entry.delete()

    def test_the_entry_survives_the_attempt(self):
        entry = self.make_entry()
        try:
            entry.delete()
        except PermissionError:
            pass
        self.assertTrue(AuditLog.objects.filter(pk=entry.pk).exists())


class FinancialAuditTrailTests(TestCase):
    """Les opérations financières laissent une trace nominative."""

    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.year = make_year(cls.school)
        cls.classroom = make_classroom(cls.school, "CP")
        make_fee_schedule(cls.school, cls.classroom, cls.year)
        cls.student = make_student(cls.school, cls.classroom)
        cls.category = make_category(cls.school, "RENT", "LOCATIONS DE BÂTIMENTS")
        cls.admin = make_user(cls.school, Role.ADMIN, "admin@test.sn")

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def payment_payload(self, tuition=15_000):
        return {
            "student": self.student.pk, "year": self.year.pk,
            "period": "2025-10-31", "tuition": tuition,
        }

    def entries_for(self, entity, action=None):
        queryset = AuditLog.objects.filter(entity=entity)
        if action:
            queryset = queryset.filter(action=action)
        return queryset

    def test_creating_a_payment_is_recorded(self):
        response = self.client.post(
            "/api/monthly-payments/", self.payment_payload(), format="json"
        )
        self.assertEqual(response.status_code, 201)

        entry = self.entries_for("MonthlyPayment", AuditLog.Action.CREATE).get()
        self.assertEqual(entry.user, self.admin)
        # Le libellé figé doit rester exploitable même si le compte disparaît,
        # d'où l'email plutôt qu'un simple nom d'affichage.
        self.assertIn("admin@test.sn", entry.user_label)
        self.assertEqual(entry.entity_id, str(response.data["id"]))
        self.assertEqual(entry.school, self.school)

    def test_the_trace_survives_the_deletion_of_its_author(self):
        self.client.post("/api/monthly-payments/", self.payment_payload(), format="json")
        entry = self.entries_for("MonthlyPayment", AuditLog.Action.CREATE).get()

        self.admin.delete()
        entry.refresh_from_db()
        self.assertIsNone(entry.user, "La clé étrangère doit passer à NULL.")
        self.assertIn("admin@test.sn", entry.user_label, "L'identité doit rester lisible.")

    def test_updating_a_payment_records_both_states(self):
        """Le montant d'origine doit rester lisible après correction."""
        created = self.client.post(
            "/api/monthly-payments/", self.payment_payload(15_000), format="json"
        )
        self.client.patch(
            f"/api/monthly-payments/{created.data['id']}/", {"tuition": 5_000}, format="json"
        )

        entry = self.entries_for("MonthlyPayment", AuditLog.Action.UPDATE).get()
        self.assertEqual(entry.before["tuition"], 15_000)
        self.assertEqual(entry.after["tuition"], 5_000)

    def test_deleting_a_payment_is_recorded_with_its_last_state(self):
        created = self.client.post(
            "/api/monthly-payments/", self.payment_payload(15_000), format="json"
        )
        payment_id = created.data["id"]
        response = self.client.delete(f"/api/monthly-payments/{payment_id}/")
        self.assertEqual(response.status_code, 204)

        entry = self.entries_for("MonthlyPayment", AuditLog.Action.DELETE).get()
        self.assertEqual(entry.entity_id, str(payment_id))
        # L'état d'avant suppression doit subsister, sinon la trace est inexploitable.
        self.assertEqual(entry.before["tuition"], 15_000)
        with tenant_context(self.school):
            self.assertFalse(MonthlyPayment.objects.filter(pk=payment_id).exists())

    def test_expense_lifecycle_is_recorded(self):
        created = self.client.post(
            "/api/expenses/",
            {
                "year": self.year.pk, "operation_date": "2025-11-10",
                "label": "Loyer novembre", "amount": 450_000, "category": self.category.pk,
            },
            format="json",
        )
        self.assertEqual(created.status_code, 201)
        self.client.delete(f"/api/expenses/{created.data['id']}/")

        actions = set(self.entries_for("Expense").values_list("action", flat=True))
        self.assertEqual(actions, {AuditLog.Action.CREATE, AuditLog.Action.DELETE})

    def test_sensitive_fields_are_not_copied_into_the_log(self):
        """Le journal ne doit pas devenir une seconde base de données personnelles."""
        response = self.client.post(
            "/api/teachers/",
            {"first_name": "Ousmane", "last_name": "Bodian", "cni": "1234567890123"},
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        entry = self.entries_for("Teacher").get()
        self.assertNotIn("cni", entry.after)

    def test_login_is_recorded(self):
        client = APIClient()
        response = client.post(
            "/api/auth/login/",
            {"email": "admin@test.sn", "password": "TestPassw0rd!"},
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(AuditLog.objects.filter(action=AuditLog.Action.LOGIN).exists())

    def test_export_is_recorded(self):
        response = self.client.get("/api/exports/bilan.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(AuditLog.objects.filter(action=AuditLog.Action.EXPORT).exists())

    def test_audit_log_is_read_only_through_the_api(self):
        AuditLog.objects.create(
            school=self.school, user=self.admin, user_label=str(self.admin),
            action=AuditLog.Action.CREATE, entity="Expense", entity_id="1",
        )
        listing = self.client.get("/api/audit-logs/")
        self.assertEqual(listing.status_code, 200)
        self.assertGreaterEqual(listing.data["count"], 1)

        entry_id = listing.data["results"][0]["id"]
        # 403 (la matrice n'accorde que la lecture) ou 405 (méthode non routée) :
        # les deux sont des refus corrects, seul l'aboutissement serait un défaut.
        for response in (
            self.client.delete(f"/api/audit-logs/{entry_id}/"),
            self.client.patch(
                f"/api/audit-logs/{entry_id}/", {"action": "DELETE"}, format="json"
            ),
        ):
            self.assertIn(response.status_code, (403, 405))
        self.assertTrue(AuditLog.objects.filter(pk=entry_id).exists())


class ExportTests(TestCase):
    """Les exports produisent bien un fichier exploitable."""

    @classmethod
    def setUpTestData(cls):
        cls.school = make_school()
        cls.year = make_year(cls.school)
        cls.classroom = make_classroom(cls.school, "CP")
        make_fee_schedule(cls.school, cls.classroom, cls.year)
        make_student(cls.school, cls.classroom)
        cls.admin = make_user(cls.school, Role.ADMIN, "admin@test.sn")

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(self.admin)

    def test_xlsx_exports_are_valid_workbooks(self):
        import io

        from openpyxl import load_workbook

        for report in ("bilan", "encais", "students"):
            response = self.client.get(f"/api/exports/{report}.xlsx")
            self.assertEqual(response.status_code, 200, f"Export {report} en échec.")
            self.assertIn("attachment;", response["Content-Disposition"])
            workbook = load_workbook(io.BytesIO(response.content))
            self.assertGreater(workbook.active.max_row, 1, f"Export {report} vide.")

    def test_pdf_export_is_a_valid_document(self):
        response = self.client.get("/api/exports/bilan.pdf")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_unknown_export_is_rejected(self):
        self.assertEqual(self.client.get("/api/exports/inexistant.xlsx").status_code, 404)
